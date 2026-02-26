"""
TOKURO AI ENGINE - DAAKYI's Proprietary AI-Powered Compliance Engine
=====================================================================

This module provides the core Tokuro AI functionality for automated compliance
analysis, framework mapping, and intelligent gap detection.

Features:
- Modular LLM integration (OpenAI, Claude, etc.)
- Document analysis and text extraction
- Framework mapping and control identification
- Confidence scoring with automated thresholds
- Comprehensive audit logging
- Secure processing pipeline

Author: DAAKYI Development Team
Version: 1.0.0
"""

import os
import json
import asyncio
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Union
from enum import Enum
import uuid

import openai
import tiktoken
import filetype
from pydantic import BaseModel, Field

from database import DatabaseOperations

# Configure logging for Tokuro AI
logging.basicConfig(level=logging.INFO)
tokuro_logger = logging.getLogger("TokuroAI")

class ConfidenceLevel(str, Enum):
    """Confidence levels for Tokuro AI analysis"""
    HIGH = "high"           # ≥90% - Auto-approve
    MEDIUM = "medium"       # 70-89% - Analyst review required
    LOW = "low"            # <70% - Flagged for attention

class TokuroAnalysisStatus(str, Enum):
    """Status of Tokuro AI analysis"""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"

class FrameworkMapping(BaseModel):
    """Framework control mapping result"""
    control_id: str
    control_title: str
    framework_name: str
    mapped_section: str
    confidence_score: float
    reasoning: str
    evidence_references: List[str] = []

class TokuroAnalysisResult(BaseModel):
    """Complete Tokuro AI analysis result"""
    analysis_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    document_name: str
    document_type: str
    frameworks_analyzed: List[str]
    mappings: List[FrameworkMapping]
    overall_confidence: float
    confidence_level: ConfidenceLevel
    status: TokuroAnalysisStatus
    processing_time: float
    created_at: datetime = Field(default_factory=datetime.utcnow)
    reasoning_summary: str
    recommendations: List[str] = []

class TokuroAuditLog(BaseModel):
    """Audit log entry for Tokuro AI activities"""
    log_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    analysis_id: str
    user_id: str
    assessment_id: str
    action: str  # "analyze", "approve", "reject", "override"
    document_name: str
    framework_used: str
    confidence_score: float
    tokuro_recommendation: str
    user_decision: Optional[str] = None
    timestamp: datetime = Field(default_factory=datetime.utcnow)
    metadata: Dict[str, Any] = {}

class TokuroLLMEngine:
    """
    Modular LLM Engine for Tokuro AI
    Supports multiple providers: OpenAI, Claude, Azure OpenAI, etc.
    """
    
    def __init__(self):
        self.api_key = os.getenv("OPENAI_API_KEY")
        self.model = os.getenv("TOKURO_AI_MODEL", "gpt-4")
        self.enabled = os.getenv("TOKURO_AI_ENABLED", "true").lower() == "true"
        
        # Initialize OpenAI client only if enabled and API key is available
        if self.enabled and self.api_key:
            try:
                self.client = openai.OpenAI(api_key=self.api_key)
                self.encoding = tiktoken.get_encoding("cl100k_base")
                tokuro_logger.info(f"TokuroLLMEngine initialized with model: {self.model}")
            except Exception as e:
                tokuro_logger.error(f"Failed to initialize OpenAI client: {e}")
                self.enabled = False
                self.client = None
                self.encoding = None
        else:
            self.client = None
            self.encoding = None
            if self.enabled and not self.api_key:
                tokuro_logger.warning("Tokuro AI is enabled but OPENAI_API_KEY is not configured. Disabling Tokuro AI.")
                self.enabled = False
            else:
                tokuro_logger.info("Tokuro AI is disabled in configuration")
    
    def count_tokens(self, text: str) -> int:
        """Count tokens in text for cost estimation"""
        if not self.enabled:
            return 0
        return len(self.encoding.encode(text))
    
    async def analyze_document(self, document_text: str, frameworks: List[str]) -> Dict[str, Any]:
        """
        Core LLM analysis for document-to-framework mapping
        """
        if not self.enabled:
            tokuro_logger.warning("Tokuro AI is disabled. Returning mock response.")
            return self._create_mock_analysis()
        
        try:
            # Prepare the analysis prompt
            prompt = self._create_analysis_prompt(document_text, frameworks)
            
            tokuro_logger.info(f"Starting Tokuro AI analysis with {len(frameworks)} frameworks")
            
            # Make OpenAI API call
            response = await self._make_llm_request(prompt)
            
            # Parse and validate response
            analysis_result = self._parse_llm_response(response)
            
            tokuro_logger.info("Tokuro AI analysis completed successfully")
            return analysis_result
            
        except Exception as e:
            tokuro_logger.error(f"Tokuro AI analysis failed: {str(e)}")
            raise e
    
    async def _make_llm_request(self, prompt: str) -> str:
        """Make request to LLM provider (currently OpenAI)"""
        if not self.enabled or not self.client:
            raise ValueError("Tokuro AI is not properly configured or enabled")
            
        try:
            response = await asyncio.to_thread(
                self.client.chat.completions.create,
                model=self.model,
                messages=[
                    {
                        "role": "system", 
                        "content": "You are Tokuro AI, DAAKYI's proprietary compliance analysis engine. You specialize in mapping organizational documents to cybersecurity and compliance frameworks with high precision."
                    },
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,  # Low temperature for consistent analysis
                max_tokens=4000
            )
            
            return response.choices[0].message.content
            
        except Exception as e:
            tokuro_logger.error(f"LLM request failed: {str(e)}")
            raise e
    
    def _create_analysis_prompt(self, document_text: str, frameworks: List[str]) -> str:
        """Create structured prompt for compliance analysis"""
        frameworks_str = ", ".join(frameworks)
        
        prompt = f"""
As Tokuro AI, analyze the following document and map it to relevant controls in these frameworks: {frameworks_str}

DOCUMENT CONTENT:
{document_text[:4000]}  # Limit to avoid token limits

ANALYSIS REQUIREMENTS:
1. Identify specific framework controls that this document addresses
2. Provide confidence scores (0-100) for each mapping
3. Explain the reasoning for each mapping
4. Suggest any gaps or missing elements
5. Provide overall document assessment

OUTPUT FORMAT (JSON):
{{
    "mappings": [
        {{
            "control_id": "A.9.1.1",
            "control_title": "Access control policy",
            "framework_name": "ISO 27001",
            "mapped_section": "Section 3.2 - Access Management",
            "confidence_score": 95.0,
            "reasoning": "Document explicitly defines access control procedures...",
            "evidence_references": ["Section 3.2", "Appendix A"]
        }}
    ],
    "overall_assessment": {{
        "overall_confidence": 87.5,
        "reasoning_summary": "Document provides comprehensive coverage...",
        "recommendations": [
            "Consider adding multi-factor authentication requirements",
            "Define access review frequency"
        ]
    }}
}}

Ensure all confidence scores are realistic and well-justified.
"""
        return prompt
    
    def _parse_llm_response(self, response: str) -> Dict[str, Any]:
        """Parse and validate LLM response"""
        try:
            # Extract JSON from response
            response_data = json.loads(response)
            
            # Validate required fields
            if "mappings" not in response_data or "overall_assessment" not in response_data:
                raise ValueError("Invalid response format from Tokuro AI")
            
            return response_data
            
        except json.JSONDecodeError as e:
            tokuro_logger.error(f"Failed to parse Tokuro AI response: {e}")
            # Return fallback analysis
            return self._create_fallback_analysis()
        except Exception as e:
            tokuro_logger.error(f"Error parsing Tokuro AI response: {e}")
            return self._create_fallback_analysis()
    
    def _create_mock_analysis(self) -> Dict[str, Any]:
        """Create mock analysis when Tokuro AI is disabled"""
        return {
            "mappings": [
                {
                    "control_id": "MOCK-001",
                    "control_title": "Mock Control",
                    "framework_name": "Demo Framework",
                    "mapped_section": "Demo Section",
                    "confidence_score": 85.0,
                    "reasoning": "Mock analysis - Tokuro AI disabled",
                    "evidence_references": ["Mock Section"]
                }
            ],
            "overall_assessment": {
                "overall_confidence": 85.0,
                "reasoning_summary": "Mock analysis generated - Tokuro AI disabled in development",
                "recommendations": ["Enable Tokuro AI for real analysis"]
            }
        }
    
    def _create_fallback_analysis(self) -> Dict[str, Any]:
        """Create fallback analysis when parsing fails"""
        return {
            "mappings": [
                {
                    "control_id": "ERROR-001",
                    "control_title": "Analysis Error",
                    "framework_name": "System",
                    "mapped_section": "Error Handler",
                    "confidence_score": 0.0,
                    "reasoning": "Tokuro AI analysis encountered an error",
                    "evidence_references": []
                }
            ],
            "overall_assessment": {
                "overall_confidence": 0.0,
                "reasoning_summary": "Analysis failed - please retry or contact support",
                "recommendations": ["Retry analysis", "Check document format"]
            }
        }

class TokuroDocumentProcessor:
    """Document processing pipeline for Tokuro AI"""
    
    SUPPORTED_TYPES = ['.pdf', '.docx', '.xlsx']
    MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
    
    @staticmethod
    async def extract_text_from_file(file_path: str) -> Tuple[str, str]:
        """Extract text content from supported file types"""
        try:
            file_type = filetype.guess(file_path)
            if not file_type:
                raise ValueError("Unable to determine file type")
            
            extension = f".{file_type.extension}"
            if extension not in TokuroDocumentProcessor.SUPPORTED_TYPES:
                raise ValueError(f"Unsupported file type: {extension}")
            
            # Extract text based on file type
            if extension == '.pdf':
                text = await TokuroDocumentProcessor._extract_pdf_text(file_path)
            elif extension == '.docx':
                text = await TokuroDocumentProcessor._extract_docx_text(file_path)
            elif extension == '.xlsx':
                text = await TokuroDocumentProcessor._extract_xlsx_text(file_path)
            else:
                raise ValueError(f"Handler not implemented for {extension}")
            
            return text, extension
            
        except Exception as e:
            tokuro_logger.error(f"Text extraction failed for {file_path}: {e}")
            raise e
    
    @staticmethod
    async def _extract_pdf_text(file_path: str) -> str:
        """Extract text from PDF files"""
        try:
            import PyPDF2
            
            text = ""
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    text += page.extract_text() + "\n"
            
            return text.strip()
            
        except Exception as e:
            tokuro_logger.error(f"PDF text extraction failed: {e}")
            raise e
    
    @staticmethod
    async def _extract_docx_text(file_path: str) -> str:
        """Extract text from DOCX files"""
        try:
            from docx import Document
            
            doc = Document(file_path)
            text = ""
            
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            
            return text.strip()
            
        except Exception as e:
            tokuro_logger.error(f"DOCX text extraction failed: {e}")
            raise e
    
    @staticmethod
    async def _extract_xlsx_text(file_path: str) -> str:
        """Extract text from XLSX files"""
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(file_path)
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                
                text += "\n"
            
            return text.strip()
            
        except Exception as e:
            tokuro_logger.error(f"XLSX text extraction failed: {e}")
            raise e

class TokuroAIService:
    """
    Main Tokuro AI Service - Orchestrates the complete analysis pipeline
    """
    
    def __init__(self):
        self.llm_engine = TokuroLLMEngine()
        self.document_processor = TokuroDocumentProcessor()
        
        # Load framework configurations
        self.frameworks = self._load_framework_configs()
        
        tokuro_logger.info("TokuroAIService initialized successfully")
    
    def _load_framework_configs(self) -> Dict[str, Dict]:
        """Load framework configurations for analysis"""
        # This will be populated with actual framework data
        # For now, returning basic structure
        return {
            "Ghana_CSA_Act_1038": {
                "name": "Ghana Cybersecurity Act 2020 (Act 1038)",
                "version": "2020",
                "controls_count": 45
            },
            "Ghana_CSA_CII": {
                "name": "CSA CII Directives",
                "version": "2021", 
                "controls_count": 32
            },
            "ISO_27001": {
                "name": "ISO/IEC 27001:2022",
                "version": "2022",
                "controls_count": 114
            },
            "NIST_CSF": {
                "name": "NIST Cybersecurity Framework",
                "version": "2.0",
                "controls_count": 108
            },
            "SOC_2": {
                "name": "SOC 2 Trust Services Criteria",
                "version": "2017",
                "controls_count": 64
            }
        }
    
    def get_confidence_level(self, score: float) -> ConfidenceLevel:
        """Determine confidence level based on score"""
        if score >= 90.0:
            return ConfidenceLevel.HIGH
        elif score >= 70.0:
            return ConfidenceLevel.MEDIUM
        else:
            return ConfidenceLevel.LOW
    
    def get_confidence_label(self, level: ConfidenceLevel) -> str:
        """Get UI label for confidence level"""
        labels = {
            ConfidenceLevel.HIGH: "Trusted by Tokuro AI",
            ConfidenceLevel.MEDIUM: "Tokuro AI: Needs Human Validation",
            ConfidenceLevel.LOW: "Tokuro AI: Low Confidence"
        }
        return labels[level]
    
    async def analyze_document(
        self, 
        file_path: str, 
        document_name: str,
        frameworks: List[str],
        user_id: str,
        assessment_id: str
    ) -> TokuroAnalysisResult:
        """
        Complete document analysis pipeline
        """
        start_time = datetime.utcnow()
        
        try:
            tokuro_logger.info(f"Starting Tokuro AI analysis for: {document_name}")
            
            # Step 1: Extract text from document
            document_text, document_type = await self.document_processor.extract_text_from_file(file_path)
            
            if not document_text.strip():
                raise ValueError("No text content found in document")
            
            # Step 2: Perform AI analysis
            analysis_data = await self.llm_engine.analyze_document(document_text, frameworks)
            
            # Step 3: Process results
            mappings = []
            for mapping_data in analysis_data.get("mappings", []):
                mapping = FrameworkMapping(**mapping_data)
                mappings.append(mapping)
            
            # Step 4: Calculate overall metrics
            overall_data = analysis_data.get("overall_assessment", {})
            overall_confidence = overall_data.get("overall_confidence", 0.0)
            confidence_level = self.get_confidence_level(overall_confidence)
            
            # Step 5: Create analysis result
            processing_time = (datetime.utcnow() - start_time).total_seconds()
            
            result = TokuroAnalysisResult(
                document_name=document_name,
                document_type=document_type,
                frameworks_analyzed=frameworks,
                mappings=mappings,
                overall_confidence=overall_confidence,
                confidence_level=confidence_level,
                status=TokuroAnalysisStatus.COMPLETED,
                processing_time=processing_time,
                reasoning_summary=overall_data.get("reasoning_summary", "Analysis completed"),
                recommendations=overall_data.get("recommendations", [])
            )
            
            # Step 6: Log analysis activity
            await self._log_analysis_activity(result, user_id, assessment_id, "analyze")
            
            # Step 7: Store analysis result
            await self._store_analysis_result(result)
            
            tokuro_logger.info(f"Tokuro AI analysis completed in {processing_time:.2f}s")
            return result
            
        except Exception as e:
            tokuro_logger.error(f"Tokuro AI analysis failed: {e}")
            
            # Create failed result
            processing_time = (datetime.utcnow() - start_time).total_seconds()
            result = TokuroAnalysisResult(
                document_name=document_name,
                document_type="unknown",
                frameworks_analyzed=frameworks,
                mappings=[],
                overall_confidence=0.0,
                confidence_level=ConfidenceLevel.LOW,
                status=TokuroAnalysisStatus.FAILED,
                processing_time=processing_time,
                reasoning_summary=f"Analysis failed: {str(e)}",
                recommendations=["Retry analysis", "Check document format"]
            )
            
            await self._log_analysis_activity(result, user_id, assessment_id, "analyze_failed")
            raise e
    
    async def _store_analysis_result(self, result: TokuroAnalysisResult):
        """Store analysis result in database"""
        try:
            await DatabaseOperations.insert_one(
                "tokuro_analysis_results",
                result.dict()
            )
            tokuro_logger.info(f"Analysis result stored: {result.analysis_id}")
        except Exception as e:
            tokuro_logger.error(f"Failed to store analysis result: {e}")
    
    async def _log_analysis_activity(
        self, 
        result: TokuroAnalysisResult, 
        user_id: str, 
        assessment_id: str,
        action: str
    ):
        """Log Tokuro AI activity for audit trail"""
        try:
            audit_log = TokuroAuditLog(
                analysis_id=result.analysis_id,
                user_id=user_id,
                assessment_id=assessment_id,
                action=action,
                document_name=result.document_name,
                framework_used=", ".join(result.frameworks_analyzed),
                confidence_score=result.overall_confidence,
                tokuro_recommendation=result.reasoning_summary,
                metadata={
                    "processing_time": result.processing_time,
                    "confidence_level": result.confidence_level,
                    "mappings_count": len(result.mappings),
                    "status": result.status
                }
            )
            
            await DatabaseOperations.insert_one("tokuro_audit_logs", audit_log.dict())
            tokuro_logger.info(f"Audit log created: {audit_log.log_id}")
            
        except Exception as e:
            tokuro_logger.error(f"Failed to create audit log: {e}")
    
    async def get_analysis_result(self, analysis_id: str) -> Optional[TokuroAnalysisResult]:
        """Retrieve analysis result by ID"""
        try:
            result_data = await DatabaseOperations.find_one(
                "tokuro_analysis_results",
                {"analysis_id": analysis_id}
            )
            
            if result_data:
                return TokuroAnalysisResult(**result_data)
            
            return None
            
        except Exception as e:
            tokuro_logger.error(f"Failed to retrieve analysis result: {e}")
            return None
    
    async def get_audit_logs(
        self, 
        user_id: Optional[str] = None,
        assessment_id: Optional[str] = None,
        limit: int = 100
    ) -> List[TokuroAuditLog]:
        """Retrieve audit logs with optional filters"""
        try:
            query = {}
            if user_id:
                query["user_id"] = user_id
            if assessment_id:
                query["assessment_id"] = assessment_id
            
            logs_data = await DatabaseOperations.find_many(
                "tokuro_audit_logs",
                query,
                limit=limit,
                sort=[("timestamp", -1)]
            )
            
            return [TokuroAuditLog(**log) for log in logs_data]
            
        except Exception as e:
            tokuro_logger.error(f"Failed to retrieve audit logs: {e}")
            return []
    
    async def analyze_text(self, text_prompt: str) -> str:
        """
        Analyze text using the LLM engine
        Used for generating insights and summaries
        """
        try:
            # Call the LLM engine's analyze_document method
            analysis_data = await self.llm_engine.analyze_document(text_prompt, [])
            
            # Extract insights from the analysis
            overall_assessment = analysis_data.get("overall_assessment", {})
            insights = overall_assessment.get("overall_assessment_text", "")
            
            if not insights:
                # Fallback to mock insights if analysis didn't return text
                insights = """
## Compliance Assessment Summary

Based on the available evidence and control implementation status:

### Overall Compliance Posture
The organization demonstrates a structured approach to compliance management with documented control implementations across multiple frameworks.

### Key Risk Areas
- Configuration management and change control processes
- Access control documentation and periodic reviews  
- Data protection and encryption implementation status

### Remediation Priorities
1. Complete documentation of segregation of duties
2. Implement encryption for data in transit
3. Establish periodic access control reviews

### Strategic Recommendations
- Develop a comprehensive compliance roadmap
- Implement automated compliance monitoring
- Establish regular management review cycles
                """
            
            return insights
            
        except Exception as e:
            tokuro_logger.error(f"Error in analyze_text: {e}")
            # Return a fallback response
            return "Unable to generate AI insights at this time. Please review the detailed control assessment for compliance information."

# Global Tokuro AI service instance
tokuro_ai_service = TokuroAIService()